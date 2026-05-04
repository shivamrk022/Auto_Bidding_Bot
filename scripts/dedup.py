# ============================================================
# scripts/dedup.py — Deduplication using Excel (openpyxl)
# ============================================================
# Tracks which posts we've already replied to, so we never
# comment twice on the same post. Also enforces daily limits.
# ============================================================

import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config.config import DB_PATH, DATA_DIR

def init_db():
    """
    Create the Excel file and headers if they don't exist yet.
    """
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(DB_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Processed Posts"
        ws.append(["post_id", "platform", "timestamp", "keyword", "comment"])
        wb.save(DB_PATH)
        print("✅ Excel database initialized.")
    else:
        print("✅ Excel database already exists.")

def is_processed(post_id: str, platform: str) -> bool:
    """
    Return True if we've already commented on this post.
    """
    if not os.path.exists(DB_PATH):
        return False
        
    wb = load_workbook(DB_PATH, read_only=True)
    ws = wb.active
    
    # Skip header
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 2 and str(row[0]) == str(post_id) and str(row[1]) == platform:
            wb.close()
            return True
            
    wb.close()
    return False

def mark_processed(post_id: str, platform: str, keyword: str = "", comment: str = ""):
    """
    Mark a post as processed so we skip it next time.
    """
    if not os.path.exists(DB_PATH):
        init_db()
        
    wb = load_workbook(DB_PATH)
    ws = wb.active
    
    timestamp = datetime.now().isoformat()
    ws.append([str(post_id), platform, timestamp, keyword, comment])
    wb.save(DB_PATH)

def get_daily_count(platform: str) -> int:
    """
    How many comments/replies have we posted today for this platform?
    """
    if not os.path.exists(DB_PATH):
        return 0
        
    today = datetime.now().date().isoformat()
    count = 0
    
    wb = load_workbook(DB_PATH, read_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 3 and str(row[1]) == platform:
            # timestamp is ISO format, e.g., '2026-05-05T04:35:32...'
            row_date = str(row[2]).split("T")[0]
            if row_date == today:
                count += 1
                
    wb.close()
    return count

def get_all_stats() -> dict:
    """
    Return a summary of all-time stats for reporting.
    """
    if not os.path.exists(DB_PATH):
        return {
            "all_time": {},
            "last_7_days": {},
            "today_linkedin": 0,
            "today_twitter": 0,
        }
        
    all_time = {}
    last_7_days = {}
    
    today = datetime.now().date()
    
    wb = load_workbook(DB_PATH, read_only=True)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) >= 3:
            platform = str(row[1])
            timestamp_str = str(row[2])
            
            # all time
            all_time[platform] = all_time.get(platform, 0) + 1
            
            # last 7 days
            try:
                row_date = datetime.fromisoformat(timestamp_str).date()
                delta = (today - row_date).days
                if 0 <= delta <= 7:
                    last_7_days[platform] = last_7_days.get(platform, 0) + 1
            except ValueError:
                pass
                
    wb.close()
    
    return {
        "all_time": all_time,
        "last_7_days": last_7_days,
        "today_linkedin": get_daily_count("linkedin"),
        "today_twitter": get_daily_count("twitter"),
    }

if __name__ == "__main__":
    init_db()
    print("Testing dedup system...")
    mark_processed("test12345", "linkedin", "need developer", "Hi, I can help!")
    print("Is processed:", is_processed("test12345", "linkedin"))   # True
    print("Not processed:", is_processed("test99999", "linkedin"))  # False
    print("Stats:", get_all_stats())
